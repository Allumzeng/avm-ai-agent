import io
import csv
import pytest
import openpyxl
from src.reporting.parser import parse_uploaded_file


def _make_excel_bytes(sheets: dict[str, list[dict]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv_bytes(rows: list[dict]) -> bytes:
    if not rows:
        return b""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def test_parse_csv_returns_rows():
    rows = [{"customer": "Alpha", "revenue": "100"}, {"customer": "Beta", "revenue": "200"}]
    result = parse_uploaded_file(_make_csv_bytes(rows), "data.csv")
    assert result["file_name"] == "data.csv"
    assert "Sheet1" in result["sheets"]
    assert len(result["sheets"]["Sheet1"]) == 2
    assert result["sheets"]["Sheet1"][0]["customer"] == "Alpha"


def test_parse_excel_returns_sheets():
    data = {"Revenue": [{"customer": "Alpha", "amount": 100}, {"customer": "Beta", "amount": 200}]}
    result = parse_uploaded_file(_make_excel_bytes(data), "report.xlsx")
    assert result["file_name"] == "report.xlsx"
    assert "Revenue" in result["sheets"]
    assert len(result["sheets"]["Revenue"]) == 2


def test_parse_excel_skips_empty_rows():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["col_a", "col_b"])
    ws.append(["val1", "val2"])
    ws.append([None, None])   # empty row — should be excluded
    ws.append(["val3", "val4"])
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_uploaded_file(buf.getvalue(), "test.xlsx")
    assert len(result["sheets"]["Data"]) == 2


def test_parse_excel_multiple_sheets():
    data = {
        "Module2": [{"activity_center": "Finance", "normal_capacity": 480}],
        "Module3": [{"activity": "Budgeting", "actual_time": 60}],
    }
    result = parse_uploaded_file(_make_excel_bytes(data), "avm.xlsx")
    assert "Module2" in result["sheets"]
    assert "Module3" in result["sheets"]


def test_parse_unsupported_extension_raises():
    with pytest.raises(ValueError, match="Unsupported"):
        parse_uploaded_file(b"junk", "report.pdf")


def test_parse_excel_deduplicates_duplicate_headers():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["amount", "amount", "name"])  # duplicate "amount"
    ws.append([100, 200, "Alpha"])
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_uploaded_file(buf.getvalue(), "dup.xlsx")
    row = result["sheets"]["Data"][0]
    assert "amount" in row
    assert "amount_1" in row
    assert row["amount"] == 100
    assert row["amount_1"] == 200
