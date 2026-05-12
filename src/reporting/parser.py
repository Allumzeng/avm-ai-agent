import csv
import io

import openpyxl


def parse_uploaded_file(file_bytes: bytes, file_name: str) -> dict:
    """Parse an Excel (.xlsx) or CSV file into a structured dict.

    Returns:
        {
            "file_name": str,
            "sheets": {
                "<sheet_name>": [{"<col>": <value>, ...}, ...]
            }
        }

    Raises:
        ValueError: if the file extension is not supported (.csv, .xlsx, .xlsm, .xltx).
    """
    ext = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
    if ext == "csv":
        return _parse_csv(file_bytes, file_name)
    if ext in ("xlsx", "xlsm", "xltx"):
        return _parse_excel(file_bytes, file_name)
    raise ValueError(f"Unsupported file type '.{ext}'. Supported: .csv, .xlsx, .xlsm, .xltx")


def _parse_excel(file_bytes: bytes, file_name: str) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers: list[str] | None = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                # Build headers, de-duplicating duplicates with a numeric suffix
                raw_headers = [
                    str(c) if c is not None else f"col_{j}"
                    for j, c in enumerate(row)
                ]
                seen: dict[str, int] = {}
                headers = []
                for h in raw_headers:
                    if h in seen:
                        seen[h] += 1
                        headers.append(f"{h}_{seen[h]}")
                    else:
                        seen[h] = 0
                        headers.append(h)
            else:
                if any(c is not None for c in row) and headers:
                    rows.append(dict(zip(headers, row)))
        if rows:
            sheets[sheet_name] = rows
    return {"file_name": file_name, "sheets": sheets}


def _parse_csv(file_bytes: bytes, file_name: str) -> dict:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = [dict(r) for r in reader if any(v.strip() for v in r.values())]
    return {"file_name": file_name, "sheets": {"Sheet1": rows}}
